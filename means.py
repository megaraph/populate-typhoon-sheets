from openpyxl import load_workbook
from pathlib import Path
from indices import VEG_INDICES, veg_index_dict

WB_NAME = "Typhoon Marilyn.xlsx"
WB_PATH = Path.cwd().joinpath(WB_NAME)

WB = load_workbook(filename=WB_NAME, data_only=True)


def main():
    before_sheets = [sheet for sheet in WB.sheetnames if sheet.startswith("B - ")]
    populate_corr_sheet(before_sheets, "Correlation Before")

    after_sheets = [sheet for sheet in WB.sheetnames if sheet.startswith("A - ")]
    populate_corr_sheet(after_sheets, "Correlation After")


def indices_results(location, index_dict):
    results = []
    for veg_index in VEG_INDICES:
        values = [row.value for row in location[index_dict[veg_index]]]
        _, *nums = values

        if nums[0] is None:
            continue

        # Perform computation on nums
        result = sum(nums) / len(nums)
        results.append(result)

    return results


def clear_sheet(sheet_name, headers):
    sheet = WB[sheet_name]
    WB.remove(sheet)

    WB.create_sheet(sheet_name)
    new_sheet = WB[sheet_name]
    new_sheet.append(headers)

    WB.save(WB_PATH)

    return new_sheet


def populate_corr_sheet(loc_sheets, corr_sheet_name):
    sample_sheet = WB[loc_sheets[0]]
    sample_sheet_titles = list(sample_sheet.rows)[
        0
    ]  # gets first row in the sample sheet

    # Get titles that are only named "Location" or any of the vegetation indices
    titles = [
        title.value
        for title in sample_sheet_titles
        if title.value in ["Location", *VEG_INDICES]
    ]

    corr_sheet = clear_sheet(sheet_name=corr_sheet_name, headers=titles)

    for sheet in loc_sheets:
        loc = WB[sheet]
        loc_name = loc["A2"].value

        index_dict = veg_index_dict()
        results = indices_results(loc, index_dict)
        row = [loc_name, *results]

        corr_sheet.append(row)

    WB.save(WB_PATH)


if __name__ == "__main__":
    main()
