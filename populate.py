from openpyxl import load_workbook
from datetime import datetime
from dateutil.relativedelta import relativedelta
from pathlib import Path
import calendar

WB_NAME = "Typhoon Marilyn.xlsx"
WB_PATH = Path.cwd().joinpath(WB_NAME)  # path to excel spreadsheet

TYPHOON_DATE = "2019-09-12"
TYPHOON_RANGE = 2

WB = load_workbook(filename=WB_NAME, data_only=True)


def populate():
    # Access the `General` sheet and its headers
    general_sheet = WB["General"]
    titles = [header.value for header in general_sheet[1]]

    # Get all sheetnames except General
    general_index = WB.sheetnames.index("General")
    all_sheets = WB.sheetnames[general_index + 1 :]

    # Clear all sheets except General
    if not len(all_sheets) == 0:
        clear_sheets(all_sheets)

    headers = [title.lower() for title in titles]
    date_index = headers.index("date")

    # Convert typhoon date from a string into a datetime object
    typhoon_date = datetime(*[int(num) for num in TYPHOON_DATE.split("-")])

    # Typhoon Date Lower and Upper Bounds (+/- 2 months)
    date_before = typhoon_date - relativedelta(months=TYPHOON_RANGE)
    date_lower = datetime(date_before.year, date_before.month, 1)

    date_after = typhoon_date + relativedelta(months=TYPHOON_RANGE)
    month_last_day = calendar.monthrange(date_after.year, date_after.month)[1]
    date_upper = datetime(date_after.year, date_after.month, month_last_day)

    # before and after datasets
    before = []
    after = []

    # Sort out images whether they were imaged BEFORE or AFTER the typhoon
    for row in general_sheet.iter_rows(min_row=2, values_only=True):
        date = row[date_index]

        if date is None:
            continue

        # Before the typhoon
        if date <= typhoon_date and date >= date_lower:
            before.append(row)

        # After the typhoon
        if date > typhoon_date and date <= date_upper:
            after.append(row)

    # Populate before sheets
    append_rows(phase="before", rows=before, headers=titles)

    # Populate after sheets
    append_rows(phase="after", rows=after, headers=titles)


def clear_sheets(all_sheets):
    for sheet_name in all_sheets:
        sheet = WB[sheet_name]
        WB.remove(sheet)
        WB.create_sheet(sheet_name)
    WB.save(WB_PATH)


def append_rows(phase, rows, headers):
    # Append data to its corresponding sheet
    for row in rows:
        if phase.lower() == "before":
            sheet_name = f"B - {row[0]}"

        if phase.lower() == "after":
            sheet_name = f"A - {row[0]}"

        if not sheet_name in WB.sheetnames:
            WB.create_sheet(sheet_name)
            WB.save(WB_PATH)

        if len(list(WB[sheet_name].rows)) == 0:
            WB[sheet_name].append(headers)

        WB[sheet_name].append(row)

    WB.save(WB_PATH)


if __name__ == "__main__":
    populate()
